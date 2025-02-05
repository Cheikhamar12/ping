import pandas as pd

from openpyxl import load_workbook

from openpyxl.styles import Alignment

import pandas as pd

from datetime import datetime, timedelta

from statsmodels.tsa.statespace.sarimax import SARIMAX

import matplotlib.pyplot as plt

from sklearn.metrics import mean_absolute_error

model_path = os.path.join(models_dir, 'sarima.pkl') 


# Définir le fichier d'entrée et de sortie
input_file = 'PRESENCE_2024.xlsx'  
year = 2024  # Modifier selon l'année traitée
output_file = 'Total_Presents_Final.csv'

# Fonction pour traiter un fichier et extraire les données
def process_file(file_path, year):
    """ Extrait les présences par semaine et les retourne sous forme de DataFrame """
    excel_data = pd.ExcelFile(file_path)
    week_sheets = sorted([sheet for sheet in excel_data.sheet_names if sheet.startswith('S')])
    results = []

    for sheet in week_sheets:
        data = pd.read_excel(file_path, sheet_name=sheet, header=None)
        row_index = data.apply(lambda row: row.astype(str).str.contains("Total présents", case=False, na=False).any(), axis=1)
        row_indices = row_index[row_index].index.tolist()

        if row_indices:
            target_row = data.iloc[row_indices[0]]
            row_sum = pd.to_numeric(target_row, errors='coerce').sum()

            
    return pd.DataFrame(results)

# Traiter le fichier actuel
new_data = process_file(input_file, year)

# Charger l'ancien fichier s'il existe, sinon créer un nouveau DataFrame
if os.path.exists(output_file):
    existing_data = pd.read_excel(output_file)
    final_df = pd.concat([existing_data, new_data], ignore_index=True)
else:
    final_df = new_data

# Sauvegarder le fichier mis à jour
final_df.to_csv(output_file, index=False)

# Fusionner les cellules pour la colonne "Année"
wb = load_workbook(output_file)
ws = wb.active

# Fusionner les cellules de la colonne "Année"
current_row = 2  # Ligne de départ (car ligne 1 = en-têtes)
for year in final_df['Annee'].unique():
    year_rows = final_df[final_df['Annee'] == year]
    start_row = current_row
    end_row = current_row + len(year_rows) - 1

    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
    cell = ws.cell(row=start_row, column=1)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    current_row = end_row + 1

# Sauvegarder le fichier avec fusion des cellules
wb.save(output_file)

print(f"Les données ont été ajoutées à {output_file} et la fusion des cellules a été appliquée.")